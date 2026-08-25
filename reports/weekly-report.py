#!/usr/bin/env python3
"""
Duxbury Trading Post — weekly inventory report.

Reads the newest Shopify products export on the Desktop plus the purchase log,
and reports what Card Ladder can't: what's sitting, what's priced under water,
and what you own but have never listed.

Builds its own history — Shopify's export has no "created at" column, so the
first run records a baseline and staleness becomes meaningful from then on.

    python3 weekly-report.py
"""
import csv, json, glob, os, re, sys
from datetime import date, datetime

HERE     = os.path.dirname(os.path.abspath(__file__))
HISTORY  = os.path.join(HERE, 'inventory-history.json')
LOG      = os.path.join(HERE, '..', 'whatnot', 'purchase-log.xlsx')

# HERE first: launchd cannot read ~/Desktop (TCC, no Full Disk Access), so a
# glob there comes back empty from a scheduled run even when the file is
# plainly sitting on the Desktop. Kept in the list for hand-runs.
SEARCH_DIRS = [HERE, os.path.expanduser('~/Desktop'), os.path.expanduser('~/Downloads')]

FEE, PER_LOW, PER_HIGH, ESE, GA = 0.1325, 0.30, 0.40, 0.78, 6.07
STALE_DAYS = 30

def net(p):
    return p*(1-FEE) - (PER_LOW if p <= 10 else PER_HIGH) - (ESE if p <= 20 else GA)

def break_even(cost):
    x = 0.01
    while x < 20000:
        if net(x) >= cost: return round(x + 0.004, 2)
        x = round(x + 0.01, 2)
    return None

def newest_export():
    files = []
    for d in SEARCH_DIRS:
        try:
            files += glob.glob(os.path.join(d, 'products_export*.csv'))
        except OSError:
            continue
    if not files:
        sys.exit('No products_export*.csv found in %s. Export from Shopify first.'
                 % ' or '.join(SEARCH_DIRS))
    return max(files, key=os.path.getmtime)

def load_shopify(path):
    rows = list(csv.DictReader(open(path)))
    hk = [k for k in rows[0].keys() if k.endswith('Handle')][0]
    def f(r, k):
        try: return float(r.get(k) or 0)
        except: return 0.0
    def bc(r): return (r.get('Variant Barcode') or '').strip().lstrip("'")
    out = []
    for r in rows:
        if not r.get('Title', '').strip(): continue
        out.append({
            'handle': r[hk], 'title': r['Title'].strip(), 'status': (r.get('Status') or '').lower(),
            'price': f(r, 'Variant Price'), 'cost': f(r, 'Cost per item'),
            'qty': f(r, 'Variant Inventory Qty') or 1, 'barcode': bc(r),
        })
    return out

def load_unlisted():
    try:
        import openpyxl
    except ImportError:
        return []
    if not os.path.exists(LOG): return []
    P = openpyxl.load_workbook(LOG, data_only=True)['Purchase log']
    out = []
    for i in range(2, P.max_row + 1):
        card = str(P.cell(i, 3).value or '').strip()
        paid = P.cell(i, 4).value
        listed = str(P.cell(i, 7).value or '').strip().lower()
        status = str(P.cell(i, 13).value or '')
        if not card or not isinstance(paid, (int, float)) or listed == 'yes': continue
        if re.search(r'sold|binned|archived', status, re.I): continue
        in_transit = bool(re.search(r'in transit|awaiting|arriv|break', status + card, re.I))
        is_box     = bool(re.search(r'sealed|opened', card, re.I))
        out.append({'row': i, 'paid': paid, 'card': card, 'transit': in_transit, 'box': is_box})
    return out

def main():
    today = date.today().isoformat()
    export = newest_export()
    cards  = load_shopify(export)
    active = [c for c in cards if c['status'] == 'active']

    hist = json.load(open(HISTORY)) if os.path.exists(HISTORY) else {'first_seen': {}, 'snapshots': {}}
    fs = hist.setdefault('first_seen', {})
    prev = hist.get('snapshots', {}).get(max(hist['snapshots'].keys()), {}) if hist.get('snapshots') else {}

    snap = {}
    for c in active:
        key = c['barcode'] or c['handle']
        fs.setdefault(key, today)
        snap[key] = {'price': c['price'], 'title': c['title']}
    hist.setdefault('snapshots', {})[today] = snap
    json.dump(hist, open(HISTORY, 'w'), indent=1)

    age = (datetime.now() - datetime.fromtimestamp(os.path.getmtime(export))).days
    print(f'DUXBURY TRADING POST — inventory report  {today}')
    print(f'source: {os.path.basename(export)}  ({age}d old)')
    if age >= 7:
        print()
        print('  !! EXPORT IS ' + str(age) + ' DAYS OLD — costs, prices and statuses below may be wrong.')
        print('  !! Shopify -> Products -> Export -> All products -> plain CSV, then re-run.')
    print()

    tot_cost = sum(c['cost']*c['qty'] for c in active)
    tot_net  = sum(net(c['price'])*c['qty'] for c in active if c['price'] > 0)
    print(f'{len(active)} active listings   cost ${tot_cost:,.2f}   net at list ${tot_net:,.2f}   ({tot_net-tot_cost:+,.2f})\n')

    # 1. below break-even
    under = []
    for c in active:
        if c['cost'] <= 0 or c['price'] <= 0: continue
        pl = net(c['price']) - c['cost']
        if pl < 0: under.append((pl, c))
    under.sort(key=lambda x: x[0])
    print(f'── PRICED BELOW BREAK-EVEN — {len(under)} cards, ${sum(-x[0] for x in under):,.2f} of embedded loss')
    for pl, c in under[:12]:
        print(f'   {pl:+9.2f}   list ${c["price"]:7.2f}  cost ${c["cost"]:7.2f}  BE ${break_even(c["cost"]):7.2f}  {c["barcode"]:10} {c["title"][:44]}')
    if len(under) > 12: print(f'   ... and {len(under)-12} more')

    # 2. stale
    stale = []
    for c in active:
        key = c['barcode'] or c['handle']
        d = (date.today() - date.fromisoformat(fs[key])).days
        if d >= STALE_DAYS: stale.append((d, c))
    stale.sort(key=lambda x: -x[1]['price'])
    print(f'\n── SITTING {STALE_DAYS}+ DAYS — {len(stale)} cards')
    if not stale:
        print(f'   (history starts {today} — staleness becomes meaningful after {STALE_DAYS} days of runs)')
    for d, c in stale[:12]:
        print(f'   {d:4d}d   list ${c["price"]:7.2f}  cost ${c["cost"]:7.2f}  {c["barcode"]:10} {c["title"][:46]}')

    # 3. price changes since last run
    moved = []
    for k, v in snap.items():
        if k in prev and abs(prev[k]['price'] - v['price']) > 0.005:
            moved.append((v['price'] - prev[k]['price'], prev[k]['price'], v['price'], v['title']))
    if moved:
        moved.sort(key=lambda x: x[0])
        print(f'\n── PRICE CHANGES SINCE LAST RUN — {len(moved)}')
        for d, o, n, t in moved[:12]:
            print(f'   {d:+8.2f}   ${o:7.2f} -> ${n:7.2f}   {t[:50]}')

    # 4. owned but never listed
    unl = load_unlisted()
    inhand  = [u for u in unl if not u['transit'] and not u['box']]
    transit = [u for u in unl if u['transit']]
    boxes   = [u for u in unl if u['box']]
    print(f'\n── OWNED BUT NOT LISTED')
    print(f'   in hand      {len(inhand):3d} rows   ${sum(u["paid"] for u in inhand):9,.2f}')
    print(f'   in transit   {len(transit):3d} rows   ${sum(u["paid"] for u in transit):9,.2f}')
    print(f'   sealed/boxes {len(boxes):3d} rows   ${sum(u["paid"] for u in boxes):9,.2f}')
    for u in sorted(inhand, key=lambda x: -x['paid'])[:10]:
        print(f'      row {u["row"]:3d}  ${u["paid"]:8.2f}  {u["card"][:56]}')

if __name__ == '__main__':
    main()
